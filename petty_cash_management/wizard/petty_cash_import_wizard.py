from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
import openpyxl
import io

class PettyCashImportWizard(models.TransientModel):
    _name = 'petty.cash.import.wizard'
    _description = 'Import Petty Cash Lines from Excel'

    file = fields.Binary("Excel File", required=True)
    filename = fields.Char()
    petty_cash_id = fields.Many2one('petty.cash', required=True)

    def action_import(self):
        if not self.file:
            raise UserError("Please upload an Excel file.")

        # ----------------------------------------
        # Decode and load workbook correctly
        # ----------------------------------------
        try:
            data = base64.b64decode(self.file)
            workbook = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError("Invalid file format. Please upload a valid .xlsx file.") from e

        # Expected columns
        COL_DATE = 0
        COL_DESCRIPTION = 1
        COL_INVOICE = 2
        COL_AMOUNT = 3
        COL_VAT_APPLICABLE = 4
        COL_CATEGORY = 5
        COL_SUPPLIER = 6
        COL_PO = 7
        COL_MR = 8
        COL_ZONE = 9

        line_model = self.env['petty.cash.line']

        for row in sheet.iter_rows(min_row=2):  # skip header
            date = row[COL_DATE].value
            description = row[COL_DESCRIPTION].value
            invoice = row[COL_INVOICE].value
            amount = row[COL_AMOUNT].value
            vat_app = row[COL_VAT_APPLICABLE].value
            category_name = row[COL_CATEGORY].value

            supplier = row[COL_SUPPLIER].value if len(row) > COL_SUPPLIER else False
            po = row[COL_PO].value if len(row) > COL_PO else False
            mr = row[COL_MR].value if len(row) > COL_MR else False
            zone = row[COL_ZONE].value if len(row) > COL_ZONE else False

            if not category_name:
                raise UserError("Category is missing in one of the rows.")

            category = self.env['petty.cash.category'].search([('name', '=', category_name)], limit=1)
            if not category:
                raise UserError(f"Category '{category_name}' not found in system.")

            # Convert VAT to boolean
            vat_bool = False
            if vat_app:
                vat_str = str(vat_app).strip().lower()
                vat_bool = vat_str in ['yes', 'y', 'true', '1', '15%', 'vat']

            # Create line
            line_model.create({
                'petty_cash_id': self.petty_cash_id.id,
                'date': date,
                'description': description,
                'invoice_number': invoice,
                'amount_before_vat': amount,
                'vat_applicable': vat_bool,
                'category_id': category.id,
                'supplier': supplier,
                'po_number': po,
                'mr_number': mr,
                'zone': zone,
            })

        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }
        #template
    template_file = fields.Binary("Template File")
    template_filename = fields.Char(default="PettyCashTemplate.xlsx")

    def action_download_template(self):
        import openpyxl
        from openpyxl.styles import Font

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Petty Cash Template"

        headers = [
            'date',
            'description',
            'invoice_number',
            'amount_before_vat',
            'vat_applicable',
            'category_name',
            'supplier',
            'po_number',
            'mr_number',
            'zone',
        ]
        sheet.append(headers)

        # Make header bold
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        # Optional example row
        sheet.append([
            '2025-01-01', 'Fuel purchase', 'INV123', 100.0, 'Yes',
            'Transportation', 'Gas Station', 'PO445', 'MR112', 'Zone A'
        ])

        # Convert to binary
        import io, base64
        stream = io.BytesIO()
        workbook.save(stream)
        file_data = base64.b64encode(stream.getvalue())

        self.template_file = file_data  # store file temporarily

        # Download
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/?model=petty.cash.import.wizard&id={self.id}&field=template_file&filename={self.template_filename}&download=true",
            'target': 'self',
        }

